from openpyxl import load_workbook
import pandas as pd


def make_dict(workbook):
    rows = 1
    columns = 1
    wb = load_workbook(workbook, read_only=True)
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    column_count = sheet.max_column
    listfolders = []
    while rows <= row_count:
        while columns <= column_count:
            name = sheet.cell(row=rows, column=columns).value
            df = pd.read_excel(workbook)  # can also index sheet by name or fetch all sheets
            mylist = df[name].tolist()
            listfolders.append(mylist)
            columns += 1
        rows += 1
    num = 0
    new_list = []
    while num < len(listfolders):
        data = listfolders[num]
        new_list.append({data[0]: data[1:]})
        num += 1
    return new_list


def duplicates(lst, item):
    lst = (list(lst.values())[0])
    return [i for i, x in enumerate(lst) if x == item]


def lecturer_subject(data, subject):
    lst = (list(data[0].values())[0].index(subject))
    lecturer = []
    for info in data:
        for name in info:
            if (list(info.values())[0][lst]) == 1:
                lecturer.append(name)
    return lecturer


def lecturer_free(data, name, time):
    lst = []
    for item in data:
        for key in item.keys():
            lst.append(key)
    no = lst.index(name)

    free = list(data[no].values())[0][time]
    if free == 1:
        return True
    else:
        return False


def class_free(data, time):
    if list(data[1].values())[0][time] == 1:
        return True
    else:
        return False


def lecturer_available_time(Lecture_Hours, Lecturer_Expertise, Lecturer_Free, subject, time):
    subject_list = []
    for item in Lecture_Hours:
        list_of_lists = (list(item.keys()))
        for x in list_of_lists:
            subject_list.append(x)
    subject_index = subject_list.index(subject)
    hours = list(Lecture_Hours[subject_index].values())[0][0]
    lst = []
    if hours > 0:
        for names in (lecturer_subject(Lecturer_Expertise, subject)):
            if lecturer_free(Lecturer_Free, names, time) == True:
                print(names + "is available at this time.")
                lst.append(names)
    if len(lst) == 0:
        print("No lecturers available at this time.")
        return ["No lecturers available at this time for this subject."]
    return lst

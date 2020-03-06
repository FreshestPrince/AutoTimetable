from openpyxl import load_workbook
import pandas as pd
from operator import itemgetter
from xlwt import Workbook
import random
from random import randrange


def make_dict(workbook):
    rows = 1
    columns = 1
    wb = load_workbook(workbook, read_only=True)
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    column_count = sheet.max_column - 1
    listfolders = []
    while rows <= row_count:
        while columns <= column_count:
            name = sheet.cell(row=rows, column=columns).value
            df = pd.read_excel(workbook)  # can also index sheet by name or fetch all sheets
            mylist = df[name].tolist()
            listfolders.append(mylist)
            columns += 1
        rows += 1
    num = 0
    new_list = []
    while num < len(listfolders):
        data = listfolders[num]
        new_list.append({data[0]: data[1:]})
        num += 1
    return new_list


def zeros(Lecturer_Free, name):
    name_lst = []
    for item in Lecturer_Free:
        list_of_lists = (list(item.keys()))
        for x in list_of_lists:
            name_lst.append(x)
    lecturer_index = name_lst.index(name)
    lst = Lecturer_Free[lecturer_index]
    lst = (list(lst.values())[0])
    zeros = len([i for i, x in enumerate(lst) if x == 0])
    return zeros


def lecturer_subject(Lecturer_Expertise, subject):
    lst = (list(Lecturer_Expertise[0].values())[0].index(subject))
    lecturer = []
    for info in Lecturer_Expertise:
        for name in info:
            if (list(info.values())[0][lst]) == 1:
                lecturer.append(name)
    return lecturer

def lunchtime(Lecturer_Free):
    data = Lecturer_Free[1:]
    for item in data:
        name = [*item][0]
        n = 3
        while n < 65:
            n_rand = randrange(0, 3)
            n = n + n_rand
            item[name][n] = 0
            n += 13
    return Lecturer_Free

def lecturer_free(Lecturer_Free, Lecturer_Expertise, name, time):

    lecturer_hours = zeros(Lecturer_Free, name)
    if lecturer_hours > 18:
        del_lecturer(Lecturer_Free, name)
        del_lecturer(Lecturer_Expertise, name)
    else:
        lst = []
        for item in Lecturer_Free:
            for key in item.keys():
                lst.append(key)
        no = lst.index(name)
        free = list(Lecturer_Free[no].values())[0][time]
        if free == 1:
            return True
        else:
            return False


def class_free(data, time):
    if list(data[1].values())[0][time] == 1:
        return True
    else:
        return False


def time_normal(number):
    while 0 < number > 13:
        number -= 13
    if number == 1:
        return 9
    elif number == 2:
        return 10
    elif number == 3:
        return 11
    elif number == 4:
        return 12
    elif number == 5:
        return 13
    elif number == 6:
        return 14
    elif number == 7:
        return 15
    elif number == 8:
        return 16
    elif number == 9:
        return 17
    elif number == 10:
        return 18
    elif number == 11:
        return 19
    elif number == 12:
        return 20
    elif number == 13:
        return 21


def time_translate(number):
    if number < 14:
        return ("Monday", time_normal(number))
    elif 27 > number > 14:
        return ("Tuesday", time_normal(number))
    elif 40 > number > 27:
        return ("Wednesday", time_normal(number))
    elif 53 > number > 40:
        return ("Thursday", time_normal(number))
    elif 66 > number > 53:
        return ("Friday", time_normal(number))
    else:
        print("Value is not within proper range")
        return 0


def lecturer_available_time(Lecture_Hours, Lecturer_Expertise, Lecturer_Free, subject, time):
    subject_list = []
    for item in Lecture_Hours:
        list_of_lists = (list(item.keys()))
        for x in list_of_lists:
            subject_list.append(x)
    subject_index = subject_list.index(subject)
    hours = list(Lecture_Hours[subject_index].values())[0][0]
    lst = []
    if hours > 0:
        for names in (lecturer_subject(Lecturer_Expertise, subject)):
            if lecturer_free(Lecturer_Free, Lecturer_Expertise, names, time) == True:
                lst.append(names)
    return lst


def check_free_classroom(data, time):
    lst = []
    free_lst = []
    number = 1
    for item in data:
        for key in item.keys():
            lst.append(key)
    while number < len(lst):
        classroom = lst[number]
        classroom_index = lst.index(classroom)
        if list(data[classroom_index].values())[0][time - 1] == 1:
            free_lst.append(classroom)
        number += 1
    return free_lst


def change_time(data, name, time):
    time_list = []
    for item in data:
        list_of_lists = (list(item.keys()))
        for x in list_of_lists:
            time_list.append(x)
    time_index = time_list.index(name)
    data[time_index][name][time] = 0


def lecturer_index(Lecturer_Free, name):
    name_lst = []
    for item in Lecturer_Free:
        list_of_lists = (list(item.keys()))
        for x in list_of_lists:
            name_lst.append(x)
    lecturer_index = name_lst.index(name)
    return lecturer_index


def del_lecturer(Lecturer_Free, name):
    name_lst = []
    for item in Lecturer_Free:
        list_of_lists = (list(item.keys()))
        for x in list_of_lists:
            name_lst.append(x)
    lecturer_index = name_lst.index(name)
    del Lecturer_Free[lecturer_index]


def find_classroom_and_lecturer(Lecture_Hours, Lecturer_Expertise, Lecturer_Free, Classroom_Free):
    lecture = []
    for value in Lecture_Hours:
        time = list(value.values())[0][0]
        subject = list(value.keys())[0]
        while time > 0:
            timeOfDay = 0
            number_of_hours = list(Classroom_Free[0].values())[0]
            while timeOfDay < len(number_of_hours):

                free_classroom = check_free_classroom(Classroom_Free, timeOfDay)
                free_lecturer = lecturer_available_time(Lecture_Hours, Lecturer_Expertise, Lecturer_Free, subject,
                                                        timeOfDay)
                if len(free_classroom) > 0 and len(free_lecturer) > 0:
                    lecturer_name = free_lecturer[random.randrange(0, len(free_lecturer))]
                    classroom_name = free_classroom[random.randrange(0, len(free_classroom))]
                    #lecturer_name = free_lecturer[0]
                    #classroom_name = free_classroom[0]
                    lecture.append([classroom_name, timeOfDay, subject,
                                    lecturer_name])
                    print(len(lecture))
                    change_time(Lecturer_Free, lecturer_name, timeOfDay)
                    change_time(Classroom_Free, classroom_name, timeOfDay)
                    timeOfDay += 1
                    time -= 1
                    if time == 0:
                        break
                    break
                else:
                    timeOfDay += 1
    return lecture


def alpha_dict_list(data, number):
    return sorted(data, key=itemgetter(number))


def make_simple_dict(data, number):
    item_list = []
    for part in data:
        item_list.append(part[number])
        item_list = list(dict.fromkeys(item_list))
    the_dict = ({y: x for x, y in enumerate(item_list, 1)})
    return the_dict


def make_excel(data, Lecturer_Free):
    wb = Workbook()
    sheet1 = wb.add_sheet('Lecturers', cell_overwrite_ok=True)
    sheet2 = wb.add_sheet("Classrooms", cell_overwrite_ok=True)
    class_dict = make_simple_dict(data, 0)
    lecturer_dict = make_simple_dict(data, 3)
    time = Lecturer_Free[0]
    for items in time:
        sheet1.write(0, 0, items)
        sheet2.write(0, 0, items)
        for hours in time[items]:
            sheet1.write(hours, 0, hours)
            sheet2.write(hours, 0, hours)
    for item in lecturer_dict:
        sheet1.write(0, lecturer_dict[item], item)
    for item in class_dict:
        sheet2.write(0, class_dict[item], item)
    for items in data:
        classroom = lecturer_dict[items[3]]
        time = items[1] + 1
        lecture = items[2]
        sheet1.write(time, classroom, lecture)
        classroom = class_dict[items[0]]
        sheet2.write(time, classroom, lecture)
    wb.save("Timetable.xls")  # Save file as an excel sheet